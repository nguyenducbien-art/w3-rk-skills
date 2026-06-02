#!/usr/bin/env python3
"""
build_xlsx.py — fill the bundled QA template (assets/template.xlsx) from a bilingual CSV.

GENERIC converter: it reads a 11-column bilingual CSV (with section-marker rows, exactly the
shape of references/example-589.csv) and renders the official Excel deliverable. It carries NO
per-screen knowledge — all data comes from the CSV + the CLI args.

Usage:
  python build_xlsx.py --csv <bilingual.csv> --screen-id 313 --screen-name "棚卸_調査明細" \
      --top "棚卸" --sub "調査明細" [--author BienND] [--today 2026-06-01] [--out <path.xlsx>]

CSV row shapes (col A decides):
  - marker row : col A = section marker text (e.g. "1. 共通確認"), cols B..K empty
  - TC row     : col A = №, cols B..K = 大分類,中分類,小分類,正常/異常,PRE(EN),前提条件(JP),
                 STEPS(EN),実施内容(JP),EXPECTED(EN),確認事項(JP)

Comments/code are English/Japanese only (never Vietnamese), per output-rules §7.
"""
import argparse, csv, datetime, re
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE = Path(__file__).resolve().parents[1] / "assets" / "template.xlsx"
TC_SHEET = "3.テストケース"
DATA_START = 18           # first data row in the template
MARKER_ROW_REF = 18       # template row to copy section-marker style/height from
TC_ROW_REF = 19           # template row to copy TC-row style/height from
LAST_COL = 49             # AW — style/clear/merge through here (result area)
N_DATA_COLS = 11          # A..K hold the test-case content

# result-area merges per TC row (2 test rounds): 結果 / 確認日 / 確認者 / 備考 ×2
RESULT_MERGES = ['M{r}:P{r}', 'Q{r}:U{r}', 'V{r}:Y{r}', 'Z{r}:AD{r}',
                 'AF{r}:AI{r}', 'AJ{r}:AN{r}', 'AO{r}:AR{r}', 'AS{r}:AW{r}']
DV_KIND   = "'0.Config'!$L$2:$L$4"   # E column: -, 正常, 異常
DV_RESULT = "'0.Config'!$A$2:$A$5"   # M/AF columns: OK, NG, PENDING, CANCEL


def is_marker(col_a):
    """A row is a section marker when col A is not an integer №."""
    s = (col_a or "").strip()
    if not s:
        return False
    try:
        int(s)
        return False
    except ValueError:
        return True


def is_date_fmt(fmt):
    """True if a number-format string is a date format (only d/m/y + separators). Used to normalize
    the template's US date formats (mm-dd-yy / m/d/yyyy) to DD/MM/YYYY everywhere."""
    s = re.sub(r"\[[^\]]*\]", "", (fmt or "").strip().lower())   # drop [$-409]-style locale tags
    return bool(s) and s not in ("general", "@") and "y" in s \
        and ("d" in s or "m" in s) and all(ch in "dmy/-. " for ch in s)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True)
    ap.add_argument("--screen-id", required=True)
    ap.add_argument("--screen-name", required=True)
    ap.add_argument("--top", default="")
    ap.add_argument("--sub", default="")
    ap.add_argument("--author", default="BienND")
    ap.add_argument("--today", default="")   # YYYY-MM-DD; default = system today
    ap.add_argument("--out", default="")
    a = ap.parse_args()

    rows = []
    with open(a.csv, encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f)
        header = next(r)                      # skip the column header row
        for row in r:
            if any((c or "").strip() for c in row):
                rows.append((row + [""] * N_DATA_COLS)[:N_DATA_COLS])

    wb = load_workbook(TEMPLATE)
    ws = wb[TC_SHEET]

    # 1.概要 — screen id + name
    ov = wb["1.概要"]
    ov["N32"] = int(a.screen_id) if a.screen_id.isdigit() else a.screen_id
    ov["N36"] = a.screen_name

    # capture reference heights + per-column styles BEFORE clearing
    section_h = ws.row_dimensions[MARKER_ROW_REF].height
    tc_h = ws.row_dimensions[TC_ROW_REF].height
    def snap(ref_row):
        out = {}
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(ref_row, c)
            out[c] = (copy(cell.font), copy(cell.fill), copy(cell.border),
                      copy(cell.alignment), cell.number_format)
        return out
    marker_style = snap(MARKER_ROW_REF)
    tc_style = snap(TC_ROW_REF)

    # unmerge everything in the data region, then clear values + heights
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= DATA_START:
            ws.unmerge_cells(str(mr))
    orig_max = ws.max_row
    for row in range(DATA_START, orig_max + 1):
        for c in range(1, LAST_COL + 1):
            ws.cell(row, c).value = None
        ws.row_dimensions[row].height = None

    def apply_style(row, style):
        for c in range(1, LAST_COL + 1):
            f, fill, border, align, numfmt = style[c]
            cell = ws.cell(row, c)
            cell.font = copy(f); cell.fill = copy(fill); cell.border = copy(border)
            cell.alignment = copy(align); cell.number_format = numfmt

    # write rows
    r = DATA_START
    tc_rows = []
    prev_num_row = None  # last TC row, for the =MAX(...)+1 auto-number
    for data in rows:
        if is_marker(data[0]):
            apply_style(r, marker_style)
            ws.cell(r, 1).value = data[0]               # marker text in col A
            ws.row_dimensions[r].height = section_h
            # The template marker style carries a medium TOP border; applied to every marker it
            # draws a heavy line above each section (incl. the 1st — the table top is the header row
            # 「№ 大分類…」, not the marker). Force a thin top border on ALL section-marker rows.
            thin = Side(style="thin", color="000000")
            for c in range(1, LAST_COL + 1):
                o = ws.cell(r, c).border
                ws.cell(r, c).border = Border(left=o.left, right=o.right, top=thin, bottom=o.bottom,
                                              diagonal=o.diagonal, outline=o.outline)
        else:
            apply_style(r, tc_style)
            # col A = auto-number formula (text marker rows are skipped by MAX)
            ws.cell(r, 1).value = f"=MAX($A${DATA_START}:A{r-1})+1" if r > DATA_START else 1
            for c in range(2, N_DATA_COLS + 1):          # B..K = content
                cell = ws.cell(r, c)
                cell.value = data[c - 1]
                old = cell.alignment
                cell.alignment = Alignment(vertical="center", wrap_text=True,
                                           horizontal=old.horizontal, indent=old.indent)
            ws.row_dimensions[r].height = tc_h
            tc_rows.append(r)
        r += 1
    last = r - 1

    # delete leftover template rows below the data
    if orig_max > last:
        ws.delete_rows(last + 1, orig_max - last)

    # result-area merges per TC row
    for tr in tc_rows:
        for m in RESULT_MERGES:
            ws.merge_cells(m.format(r=tr))

    # data validations: clear template-pinned ones, re-add only on real TC rows
    ws.data_validations.dataValidation = []
    dv_kind = DataValidation(type="list", formula1=DV_KIND, allow_blank=True)
    dv_res = DataValidation(type="list", formula1=DV_RESULT, allow_blank=True)
    dv_kind.showErrorMessage = True; dv_res.showErrorMessage = True
    ws.add_data_validation(dv_kind); ws.add_data_validation(dv_res)
    for tr in tc_rows:
        dv_kind.add(f"E{tr}")
        dv_res.add(f"M{tr}"); dv_res.add(f"AF{tr}")

    # patch the bottom border of the last TC row (template row 19 has bot=None on cols 26-35)
    if tc_rows:
        thin = Side(style="thin", color="000000")
        lr = tc_rows[-1]
        for c in range(1, LAST_COL + 1):
            o = ws.cell(lr, c).border
            ws.cell(lr, c).border = Border(left=o.left, right=o.right, top=o.top, bottom=thin,
                                           diagonal=o.diagonal, outline=o.outline)

    # 2.履歴 — author + dates. (Date display is normalized to DD/MM/YYYY by the sweep below — the
    # template ships US formats, and 3.テストケース mirrors AY1/AY2 via formula with its own format.)
    today = datetime.date.fromisoformat(a.today) if a.today else datetime.date.today()
    h = wb["2.履歴"]
    h["AO1"] = a.author; h["AY1"] = today; h["B6"] = today

    # Normalize EVERY date cell to DD/MM/YYYY across all sheets. The template ships US formats
    # (mm-dd-yy / m/d/yyyy); the 作成日/変更日 shown on 3.テストケース are formula mirrors of
    # 2.履歴!AY1/AY2 with their OWN US format, plus the 確認日 result columns — so fixing only AY1/B6
    # left the visible date wrong. An explicit DD/MM/YYYY format is locale-proof.
    for wsx in wb.worksheets:
        for row in wsx.iter_rows():
            for cell in row:
                if is_date_fmt(cell.number_format):
                    cell.number_format = "dd/mm/yyyy"

    # output path: the Excel goes in its own IT-delivery folder next to the CSV:
    #   <csv-dir>/<screen_name> (screen_id=<X>)_QAテスト/【W3…】結合テスト仕様書_<top> (<sub>) (screen_id=<X>)_ver1.0.xlsx
    if a.out:
        out = Path(a.out)
    else:
        top = a.top or a.screen_name
        fname = f"【W3 フロントエンドマイグレーション】結合テスト仕様書_{top} ({a.sub}) (screen_id={a.screen_id})_ver1.0.xlsx"
        qa_folder = f"{a.screen_name} (screen_id={a.screen_id})_QAテスト"
        out = Path(a.csv).resolve().parent / qa_folder / fname
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")
    print(f"  TC rows: {len(tc_rows)} | marker rows: {len(rows)-len(tc_rows)} | last row: {last}")


if __name__ == "__main__":
    main()
