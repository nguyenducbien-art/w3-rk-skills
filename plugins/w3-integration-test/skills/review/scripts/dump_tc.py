#!/usr/bin/env python3
"""
dump_tc.py — read a W3 integration-test-spec Excel and dump its test cases to a readable CSV
so the reviewer (Claude) can read the TC inventory without parsing the workbook by hand.

Usage:
  python dump_tc.py --xlsx "<path/結合テスト仕様書_… .xlsx>" [--out <dump.csv>]

Reads sheet `3.テストケース` (header row 16, data from row 18) → 11-column CSV with section-marker
rows preserved (col A = marker text for section rows, № for TC rows) — same shape as the gen
skill's example CSVs. Also reads `1.概要` N32 (screen_id) / N36 (screen name) and prints a summary
for the File Suitability Gate. Comments/code: English/Japanese only.
"""
import argparse, csv, re
from pathlib import Path
from openpyxl import load_workbook

TC_SHEET = "3.テストケース"
HEADER_ROW = 16
DATA_START = 18
N_COLS = 11  # A..K

# --- auto-flag heuristics (mirror the gen skill's build_csv.py verify(); text-only, screen-agnostic) ---
FORBIDDEN = ('stopLoading', 'startLoading', '$scope', '$http', 'dataBound', '.error()',
             'createEModal', 'createIModal')


def _steps(s):  # top-level step numbers in a multiline cell (1.1/1.2 → step 1)
    nums = set()
    for ln in (s or '').split('\n'):
        m = re.match(r'^\s*(\d+)', ln.strip())
        if m:
            nums.add(int(m.group(1)))
    return sorted(nums)


def _pre_crammed(pre):   # §precondition: a PRE line with a (screen_id=…) ref AND a click = crammed nav
    for ln in (pre or '').split('\n'):
        if 'screen_id=' in ln and ('click' in ln.lower() or 'クリック' in ln):
            return True
    return False


def _pre_mode_tail(pre):   # §precondition: landing line narrates mode/state (keep it terse "… is displayed.")
    low = (pre or '').lower()
    if any(k in low for k in ('opens in ', 'in edit mode', 'in new-registration mode',
                              'in new registration mode', 'empty form', 'pre-filled', 'prefilled')):
        return True
    return any(k in (pre or '') for k in ('モードで', '空フォーム'))


def _verify_sql_noprefix(cell):   # §verify-db: a post-submit verify SELECT in a STEP must be tagged [verify DB]/[DB確認]
    for ln in (cell or '').split('\n'):
        low = ln.lower()
        is_sql = 'run sql' in low or 'sqlを実行' in low
        if is_sql and 'select' in low and '[verify db]' not in low and '[db確認]' not in low:
            return True
    return False


def _crammed_subnums(cell):   # output-rules §6: 2+ dotted sub-numbers run together on one line
    for ln in (cell or '').split('\n'):
        if len(re.findall(r'\d+\.\d+', ln)) >= 2:
            return True
    return False


def _jammed_list(cell):   # §enumerate: 4+ ・ on one line = jammed field list
    for ln in (cell or '').split('\n'):
        if ln.count('・') >= 4:
            return True
    return False


def _collapsed_range(cell):   # §enumerate: a label glued to a numeric range (送状備考1〜3)
    return re.search(r'[ぁ-んァ-ヶ一-龥ー]\d+[〜～]\d+', cell or '') is not None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", default="")
    a = ap.parse_args()

    src = Path(a.xlsx)
    wb = load_workbook(src, data_only=True)

    # --- File Suitability Gate inputs ---
    sid = name = None
    if "1.概要" in wb.sheetnames:
        ov = wb["1.概要"]
        sid = ov["N32"].value
        name = ov["N36"].value
    if TC_SHEET not in wb.sheetnames:
        print(f"UNSUITABLE: no '{TC_SHEET}' sheet. Sheets: {wb.sheetnames}")
        return
    ws = wb[TC_SHEET]

    # find last content row (col A or B non-empty)
    last = DATA_START
    for r in range(DATA_START, ws.max_row + 1):
        if (ws.cell(r, 1).value not in (None, "")) or (ws.cell(r, 2).value not in (None, "")):
            last = r

    rows, tc_count, markers = [], 0, []
    for r in range(DATA_START, last + 1):
        vals = []
        for c in range(1, N_COLS + 1):
            v = ws.cell(r, c).value
            if v is None:
                vals.append("")
            elif isinstance(v, float) and v.is_integer():
                vals.append(str(int(v)))
            else:
                vals.append(str(v))
        a0, b0 = vals[0].strip(), vals[1].strip()
        is_marker = bool(a0) and not a0.isdigit() and not b0   # marker: col A text, col B empty
        if is_marker:
            markers.append(a0)
        elif b0:                               # TC row = has 大分類 (col B) — robust vs uncached № formula
            tc_count += 1
            if not a0:                         # № is a formula not cached (file never opened in Excel)
                vals[0] = str(tc_count)        # running fallback number so the reviewer can cite rows
        else:
            continue                           # blank / spacer row — skip
        rows.append(vals)

    header = ['№', '大分類', '中分類', '小分類', '正常/異常', 'PRE-CONDITION', '前提条件',
              'STEPS', '実施内容', 'EXPECTED RESULT', '確認事項']
    out = Path(a.out) if a.out else src.with_name(src.stem + "_dump.csv")
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)

    # auto-flags: run the gen verify() text heuristics over the dumped rows, tracking the section
    flags, section, n = [], "", 0
    for vals in rows:
        a0, b0 = vals[0].strip(), vals[1].strip()
        if b0 == "" and a0 and not a0.isdigit():        # marker row → remember current section
            section = a0
            continue
        if b0 == "":
            continue
        n += 1
        num, cat1, cat2, cat3, kind, pe, pj, se, sj, ee, ej = vals[:N_COLS]
        tag = (num.strip() or str(n))
        if cat3.strip() and not cat2.strip():
            flags.append(f"№{tag} {cat1}: 中分類 blank while 小分類 filled")
        if section.startswith("3.") and kind.strip() == "正常":
            flags.append(f"№{tag} {cat1}: 正常 case in バリデーション(S3) — should move to S4.1 (§s3-abnormal-only)")
        if _pre_crammed(pe) or _pre_crammed(pj):
            flags.append(f"№{tag} {cat1}: PRE line crams screen-ref + click — split into 1 action/step (§precondition)")
        if _pre_mode_tail(pe) or _pre_mode_tail(pj):
            flags.append(f"№{tag} {cat1}: PRE narrates landed-screen mode/state — keep landing terse ('… screen is displayed.'), drop 'in … mode / empty form / pre-filled' (§precondition)")
        if _verify_sql_noprefix(se) or _verify_sql_noprefix(sj):
            flags.append(f"№{tag} {cat1}: STEP runs a verify SELECT without the [verify DB] / [DB確認] prefix (§verify-db)")
        for label, cell in (('STEPS', se), ('実施内容', sj), ('EXPECTED', ee), ('確認事項', ej), ('PRE', pe), ('前提条件', pj)):
            if _crammed_subnums(cell):
                flags.append(f"№{tag} {cat1}: {label} runs 2+ sub-numbers on one line (output-rules §6)")
            if _jammed_list(cell):
                flags.append(f"№{tag} {cat1}: {label} jams 4+ fields with ・ — comma-separate (§enumerate)")
            if _collapsed_range(cell):
                flags.append(f"№{tag} {cat1}: {label} collapses a range (e.g. 送状備考1〜3) — spell out each field (§enumerate)")
        if _steps(se) != _steps(ee):
            flags.append(f"№{tag} {cat1}: STEPS{_steps(se)} vs EXPECTED-EN{_steps(ee)} step-number mismatch")
        if _steps(sj) != _steps(ej):
            flags.append(f"№{tag} {cat1}: 実施内容{_steps(sj)} vs 確認事項{_steps(ej)} step-number mismatch")
        for tok in FORBIDDEN:
            if tok in ee or tok in ej:
                flags.append(f"№{tag} {cat1}: forbidden token '{tok}' in EXPECTED")

    # §nav-button-modal (1b): the screen name (1.概要 N36) should match the user-facing page title.
    pt = None
    for vals in rows:
        if vals[1].strip() == 'ページタイトル':
            m = re.search(r'W3 mimosa\s*\|\s*(.+?)\s*["」]', (vals[9] or '') + ' ' + (vals[10] or ''))
            if m:
                pt = m.group(1).strip()
            break
    if pt and name and str(name).strip() != pt \
            and (str(name).strip() in pt or pt in str(name).strip()):
        flags.append(f"screen name (1.概要 N36='{name}') vs page title ('{pt}') look like short/long forms "
                     f"— unify to the fuller user-facing name in N36 / file / refs (§nav-button-modal)")

    print(f"screen_id (1.概要 N32): {sid}")
    print(f"screen_name (1.概要 N36): {name}")
    print(f"TC count: {tc_count} | section markers: {len(markers)} -> {markers}")
    print(f"dump CSV: {out}")
    if not sid or tc_count == 0:
        print("⚠️ SUITABILITY: missing screen_id or no TC rows — confirm this is the right file/sheet before reviewing.")
    if flags:
        print(f"\nauto-flags ({len(flags)}) — wording/structure issues to verify (heuristic, mirror gen verify()):")
        for fl in flags:
            print("  FLAG", fl)
    else:
        print("auto-flags: none (wording/structure heuristics clean)")
    print("(These are TEXT heuristics. Screen-type checks — DB-verify / concurrent / direct-URL / "
          "accordion / branch coverage — stay manual via review-checklist §3/§4/§4b.)")


if __name__ == "__main__":
    main()
